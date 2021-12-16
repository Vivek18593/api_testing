import openpyxl
import requests

#----GIVE THE PATH OF THE EXCEL FILE WHICH CONTAINS THE APIs----#
path = " "

expected_response_code = 200

#----APIs HAVING GET METHOD----#
def get_method(api,row):
    result = requests.get(api)
    #print(result.status_code)
    wb = openpyxl.load_workbook(path)
    sheet = wb["APIS"]
    if result.status_code == expected_response_code:
        sheet.cell(row,4).value = result.status_code
        sheet.cell(row,5).value = result.text
        sheet.cell(row,6).value = "PASSED"
    else:
        sheet.cell(row,4).value = result.status_code
        sheet.cell(row,5).value = result.text
        sheet.cell(row,6).value = "FAILED"
    wb.save(path)

#----APIs HAVING POST METHOD----#
def post_method(api,row):
    result = requests.post(api)
    #print(result.status_code)
    wb = openpyxl.load_workbook(path)
    sheet = wb["APIS"]
    if result.status_code == expected_response_code:
        sheet.cell(row,4).value = result.status_code
        sheet.cell(row,5).value = result.text
        sheet.cell(row,6).value = "PASSED"
    else:
        sheet.cell(row,4).value = result.status_code
        sheet.cell(row,5).value = result.text
        sheet.cell(row,6).value = "FAILED"
    wb.save(path)

#----FETCHING APIs FROM THE EXCEL FILE----#
def get_api():
    print("Test Started!")
    wb = openpyxl.load_workbook(path)
    sheet = wb["APIS"]
    rowcount = sheet.max_row
    for row in range(2,rowcount+1):
        test_method = sheet.cell(row,2).value
        #print(test_method)
        if test_method == "GET":
            test_api = sheet.cell(row,3).value
            #print(test_method,test_api)
            get_method(test_api,row)
        elif test_method == "POST":
            test_api = sheet.cell(row,3).value
            #print(test_method,test_api)
            post_method(test_api,row)


#----EXECUTING SCRIPT----#
get_api()
print("Test Completed!")



