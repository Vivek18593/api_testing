import requests, openpyxl, os
from openpyxl.styles import Font
from termcolor import colored

print('   *******************')
print('   *   API TESTING   *')
print('   *******************')

#--------EXCEL FILE--------#
current_dir = os.path.dirname(os.path.realpath(__file__))
excel_file = current_dir+'\\api_testing_result_demo.xlsx'
wb = openpyxl.load_workbook(excel_file)
sheet = wb.active
#**************************#

#-----------STATUS OF THE API-----------#
data = {'status':'false'}
json_data = ''
#***************************************#

#-----------API METHODS-----------#
response_code = [200,201,202,204]

def write_to_excel(result,row):
    if result.status_code in response_code:
        sheet.cell(row,4).value = result.status_code
        sheet.cell(row,4).font = Font(bold='True',color='000000')
        sheet.cell(row,5).value = 'PASSED'
        sheet.cell(row,5).font = Font(bold='True',color='008000')
        sheet.cell(row,6).value = result.text
        sheet.cell(row,6).font = Font(bold='True',color='000000')
        json_data = result.text
        if data['status'] in json_data:
            sheet.cell(row,7).value = 'DATA ERROR'
            sheet.cell(row,7).font = Font(bold='True',color='FF2400')
        else:
            sheet.cell(row,7).value = 'OK'
            sheet.cell(row,7).font = Font(bold='True',color='008000')
    else:
        sheet.cell(row,4).value = result.status_code
        sheet.cell(row,4).font = Font(bold='True',color='FF2400')
        sheet.cell(row,5).value = 'FAILED'
        sheet.cell(row,5).font = Font(bold='True',color='FF2400')
        sheet.cell(row,6).value = result.text
        sheet.cell(row,6).font = Font(bold='True',color='000000')

def get_method(api,row):
    result = requests.get(api)
    write_to_excel(result,row)

def post_method(api,row):
    result = requests.post(api)
    write_to_excel(result,row)

def put_method(api,row):
    result = requests.put(api)
    write_to_excel(result,row)

def patch_method(api,row):
    result = requests.patch(api)
    write_to_excel(result,row)

def delete_method(api,row):
    result = requests.delete(api)
    write_to_excel(result,row)
#**********************************#

#----------------MESSAGES----------------#
def test_result_message(row):
    api_title = sheet.cell(row,1).value
    response_code_status = sheet.cell(row,5).value
    json_data_status = sheet.cell(row,7).value
    if response_code_status == 'PASSED' and json_data_status == 'OK':
        sheet.cell(row,8).value = f'{api_title} Success'
        sheet.cell(row,8).font = Font(bold='True',color='008000')
        print(colored(api_title,'green'))
    else:
        sheet.cell(row,8).value = f'{api_title} Failed'
        sheet.cell(row,8).font = Font(bold='True',color='FF2400')
        print(colored(api_title,'red'))
#****************************************#

#-----TESTING APIS AND RECORDING RESULT TO AN EXCEL FILE-----#
def execute_api_testing():
    print('[>] Initiating API Testing!')
    rowcount = sheet.max_row
    for row in range(2,rowcount+1):
        test_method = sheet.cell(row,2).value
        if test_method == 'GET':
            test_api = sheet.cell(row,3).value
            get_method(test_api,row)
            test_result_message(row)
        elif test_method == 'POST':
            test_api = sheet.cell(row,3).value
            post_method(test_api,row)
            test_result_message(row)
        elif test_method == 'PUT':
            test_api = sheet.cell(row,3).value
            put_method(test_api,row)
            test_result_message(row)
        elif test_method == 'PATCH':
            test_api = sheet.cell(row,3).value
            patch_method(test_api,row)
            test_result_message(row)
        elif test_method == 'DELETE':
            test_api = sheet.cell(row,3).value
            delete_method(test_api,row)
            test_result_message(row)
    wb.save(excel_file)
    print('[#] Test Completed!')
#************************************************************#


#------EXECUTE------#
execute_api_testing()



